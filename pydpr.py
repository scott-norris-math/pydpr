import copy
import time
import os.path
import numpy as np
import matplotlib.pyplot as plt
from subprocess import call
from openpyxl import load_workbook
from colorama import init, Fore, Back, Style



init()


# lookup tables for common codes
termcodes = ['Jan', 'Spring', 'May', 'Summer', 'August', 'Fall', 'Fall']
degreetext = dict({'BA':'Bachelor of ARTS (no science)', 'BS':'Bachelor of Science', 'MN':'Minor'})

# Define a dictionary for grade to GPA conversion
GPAlookup = dict()
GPAlookup[' '] = 4.0
GPAlookup['NOW'] = 4.0
GPAlookup['CR'] = 4.0

GPAlookup['A']  = 4.0
GPAlookup['A-'] = 3.7
GPAlookup['B+'] = 3.3
GPAlookup['B']  = 3.0
GPAlookup['B-'] = 2.7
GPAlookup['C+'] = 2.3
GPAlookup['C']  = 2.0
GPAlookup['C-'] = 1.7
GPAlookup['D+'] = 1.3
GPAlookup['D']  = 1.0
GPAlookup['D-'] = 0.7
GPAlookup['F']  = 0.0

# courses that do not factor into GPA
GPAlookup['W'] = 0.0
GPAlookup['I'] = 0.0
GPAlookup['P'] = 0.0
GPAlookup['S'] = 0.0
GPAlookup['S-'] = 0.0
GPAlookup['NC'] = 0.0

# excluded courses for which no credit was awarded
GPAlookup['ED+'] = 1.3
GPAlookup['ED'] = 1.0
GPAlookup['ED-'] = 0.7
GPAlookup['EF'] = 0.0

# transfer courses for which no credit was awarded
GPAlookup['TD+'] = 1.3
GPAlookup['TD'] = 1.0
GPAlookup['TD-'] = 0.7
GPAlookup['TF'] = 0.0


GPAgrades  = set(['A', 'A-', 'B+', 'B', 'B-', 'C+', 'C', 'C-', 'D+', 'D', 'D-', 'F'])
catt_grades = set(['NOW', 'CR', 'A', 'A-', 'B+', 'B', 'B-', 'C+', 'C', 'C-', 'D+', 'D', 'D-', 'F', 'P', 'S', 'S-'])
cawd_grades = set(['NOW', 'CR', 'A', 'A-', 'B+', 'B', 'B-', 'C+', 'C', 'C-', 'D+', 'D', 'D-', 'P', 'S', 'S-'])
major_grades = set(['NOW', 'CR', 'A', 'A-', 'B+', 'B', 'B-', 'C+', 'C', 'C-', 'P', 'S'])
gpts_grades = set(['A', 'A-', 'B+', 'B', 'B-', 'C+', 'C', 'C-', 'D+', 'D', 'D-', 'F'])




def get_gradepoints(grade):

  if grade in GPAgrades:
    return GPAlookup[grade]
  else:
    return None




class Error(Exception):
# base class for exceptions in this module
  pass

class MissingStudentError(Error):
# raised when a student is not found in the student list

  def __init__(self, studentID, queryfile):
    self.studentID = studentID
    self.queryfile = queryfile

class MissingCoursesError(Error):
# raised when a student is not found in the coursehistory list

  def __init__(self, studentID, queryfile):
    self.studentID = studentID
    self.queryfile = queryfile

class MissingFileError(Error):
# raised when a student is not found in the coursehistory list

  def __init__(self, queryfile):
    #self.studentID = studentID
    self.queryfile = queryfile




# =====================================
#     Input / Output
# =====================================



def find_student(studentfile, fragment):

  # open the excel file
  wb = load_workbook(studentfile)
  ws = wb[wb.get_sheet_names()[0]]

  # search for matches
  matches = []
  for jj,row in enumerate(ws):
    ID    = str(row[0].value)
    lname = str(row[2].value)
    fname = str(row[3].value)
    email = str(row[6].value)

    fields = [ID, lname, fname, email]
    for kk,ff in enumerate(fields):
      pos = ff.lower().find(fragment)
      if pos > -1:
        matches.append(fields)
        break

  # if multiple matches
  if len(matches) > 1:
    for ii,match in enumerate(matches):
      print('(%1d) [%s] "%s, %s" <%s>' % (ii+1, match[0], match[1], match[2], match[3]))
    val = int(input("\nSelect a student: "))
    if val > 0 and val <= len(matches): 
      return matches[val-1][0]
    else:
      print ("Invalid entry.  Exiting.")
      exit

  # if just one match
  if len(matches) == 1:
    match = matches[0]
    print('found: [%s] "%s, %s" <%s>' % (match[0], match[1], match[2], match[3]))    
    return match[0]

  # if no matches
  print ("No matches not found.  Exiting.")
  exit







def load_student_from_query(queryfile, studentID):
  """
  This function reads $queryfile [which is assumed to be an XLSX file] looking for a given $studentID in column 0.  If the student ID is found, then demographic information is read from the row.  At the moment, the columns associated with each piece of information are hard-coded.  This could be set up in a configuration file if desired.
  """

  # read the list of students
  wb = load_workbook(queryfile)
  ws = wb[wb.get_sheet_names()[0]]

  # get desired student from the list
  record = [a for a in ws if str(a[0].value) == str(studentID)]
  if len(record) == 0:
    print("Student ID %s not found in student file %s." % (studentID, queryfile))
    raise MissingStudentError(studentID, queryfile)

  # extract a summary of the student's information
  s = Student()
  row = record[0]
  s.ID  = str(row[0].value)
  s.lname = row[2].value
  s.fname = row[3].value
  s.email = row[6].value
  s.degree = row[26].value
  s.degreecode = s.degree.split('-')[1]

  s.speccode = row[28].value
  if s.speccode != None:
    s.speccode = s.speccode.split('-')[1]

  # admit and expected graduation terms
  s.aterm = str(row[18].value)
  s.gterm = str(row[19].value)

  s.gyear  = 2000+int(s.gterm[1:3])
  s.gsess  = int(s.gterm[3])
  s.gstatus = row[20].value

  return s




def load_courses_from_query(queryfile, studentID):
  """
  This function reads $queryfile [which is assumed to be an XLSX file] looking for all rows containing a given $studentID in column 0.  Each row is assumed to represent a course taken by that student, and course details are read from the row, resulting in a list of Course() data structures.  At the moment, the columns associated with each piece of course information are hard-coded.  This could be set up in a configuration file if desired.
  """

  wb   = load_workbook(queryfile)
  ws   = wb[wb.get_sheet_names()[0]]

  #for a in ws:
  #  print a[0].value, studentID, a[0].value == studentID
  rows = [a for a in ws if str(a[0].value) == str(studentID)]

  if len(rows) == 0:
    print("Student ID %s not found in course query file %s." % (studentID, queryfile))
    raise MissingCoursesError(studentID, queryfile)

  courselist = []
  for row in rows:
    course = Course()
    course.term    = row[4].value
    course.dept    = row[5].value
    course.number  = row[6].value.strip()
    course.code    = course.dept + ' ' + str(course.number)
    course.name    = row[8].value

    course.grade   = row[9].value
    course.credits = row[10].value
    course.awarded = row[11].value
    course.type    = row[12].value

    if course.credits == 0:
      continue

    if course.grade == None:
      course.grade = 'NOW'

    courselist.append(course)

  return courselist




# =====================================
#     Utilities
# =====================================


def current_term():

  date = time.localtime()
  year = date.tm_year
  month = date.tm_mon

  term = 1000 + (year - 2000)*10
  if month <= 6:  term += 2
  if month > 6:  term += 7
  return term




def termcode_to_text(termcode):

  if termcode == None:  return None
  termcode = str(termcode)

  try:
    century = int(termcode[0])
    year2d  = int(termcode[1:3])
    term    = int(termcode[3])
    yeartxt = str(1900 + century*100 + year2d)
    termtxt = termcodes[term-1]
    return yeartxt + ' ' + termtxt
  except Exception:
    return termcode




def convert_term_name(termname):

  words = termname.split(' ')
  season = words[0]
  year   = words[1]
  syear  = int(year[2:])

  if season == 'Fall':
    newname = "%2d%2d-fall" % (syear, syear+1) 
  else:
    newname = "%2d%2d-spring" % (syear-1, syear)
  return newname



def unplanned_terms(student, coursehistory):

  gradterm = int(student.gterm)
  last_planned = int(max([c.term for c in coursehistory]))
  unplanned = (gradterm - last_planned) / 5
  return unplanned



def outside_interests(student, coursehistory, cutoff=6):

  from collections import Counter

  cox_set = set(['ACCT', 'FINA', 'ITOM', 'MKTG', 'BL', 'BLI'])

  student_department = student.degree.split('-')[0]
  dept_list = [a.dept for a in coursehistory if (a.dept != student_department and a.grade in cawd_grades)]
  dept_list = ['COX' if a in cox_set else a for a in dept_list]    

  dept_dict = Counter(dept_list)
  count_list = [(a, dept_dict[a]) for a in dept_dict]
  sorted_counts = sorted(count_list, key = lambda a: a[1], reverse=True)
  final_list = [a for a in sorted_counts if a[1] >= cutoff]
  return final_list



# =====================================
#     Essential Objects
# =====================================




class Student:
  '''
  A Student is an object for storing student information.  This includes

  - student ID
  - last name
  - first name
  - email address
  - graduation date
  - degree and specialization
  '''

  def __init__(self, ID="", lname="", fname="", email="", gcode="", degree="", subtype="", gyear="", aterm="", gterm="", degreecode="", speccode=""):

    self.ID = ID
    self.lname = lname
    self.fname = fname
    self.email = email
    self.gcode = gcode
    self.gyear = gyear
    self.aterm = aterm
    self.gterm = gterm
    #self.degree = degree
    #self.subtype = subtype
    #self.degreecode = degreecode
    #self.speccode = speccode



class Course:
  '''
  class Course()
  ----------------
  A Course is an object for storing courses that have been taken or are being taken.  Field include
  - course code
  - course title
  - term
  - grade
  - credits
  - credit status
  '''


  def __init__(self, code=None, name=None, term=None, grade=None, credits=0, status=None):

    self.dept    = None
    self.number  = None
    self.code    = code
    self.name    = name
    self.term    = term

    self.grade   = grade
    self.credits = credits
    self.awarded = None
    self.type    = None


  def credit_status(self):

    # credits attempted
    credits_attempted = self.credits if (self.grade in catt_grades) else 0.0
    credits_awarded   = self.credits if (self.grade in cawd_grades) else 0.0
    gpoints_awarded   = self.credits * GPAlookup[self.grade] if (self.grade in gpts_grades) else 0.0
    return (credits_attempted, credits_awarded, gpoints_awarded)




class Requirement:
  '''
  Requirement()
  ---------------
  The most basic building block for describing degrees, a Requirement is simply a list of courses, with a parameter specifying the minimum number of such courses that must be taken.  Examples include

  - a single, absolutely-required course       Requirement("Calc I", ["MATH 1337"], 1)
  - a certain number of courses from a list    Requirement("Two Science", ["", "", "", ...], 2)
  '''


  def __init__(self, name, courselist, mincourses=0, minhours=0, greedy=False):

    self.name       = name
    self.courselist = set(courselist)

    self.minhours   = minhours
    self.mincourses = mincourses
    self.greedy     = greedy

    self.satcourses = []
    self.satisfied  = False

    if mincourses > 0:
      self.hours_rem = 3*mincourses
    if minhours > 0:
      self.hours_rem = minhours
    
    return 


  def check(self, coursehistory, verlist=None):

    # identify satisfying courses that also satisfy verifications
    vercourses = set()
    if verlist != None:
      for ver in verlist:
        for course in ver.courselist:
          if course in self.courselist:
            vercourses.add(course)

    # first find satisfying courses that are *also* in the verification lists
    templist = []
    for course in vercourses:
      satisfying_course = [cc for cc in coursehistory if (cc.code == course and cc.credits > 0 and cc.grade in major_grades)]
      if satisfying_course != []:
        templist.append(satisfying_course[0])
    templist.sort(key = lambda course: (course.term, course.code))
    self.satcourses.extend(templist)

    # now find satisfying courses that are *not* in the verification list
    templist = []
    for course in self.courselist - vercourses:
      satisfying_course = [cc for cc in coursehistory if (cc.code == course and cc.credits > 0 and cc.grade in major_grades)]
      if satisfying_course != []:
        templist.append(satisfying_course[0])
    templist.sort(key = lambda course: (course.term, course.code))
    self.satcourses.extend(templist)

    # check satisfaction and obtain remaining hours
    if self.mincourses > 0:
      self.satisfied = (len(self.satcourses) >= self.mincourses)
      self.hours_rem = 0 if self.satisfied else 3*(self.mincourses - len(self.satcourses))

    if self.minhours > 0:
      cumhours = 0
      for course in self.satcourses:
        cumhours += course.credits
      self.satisfied = (cumhours >= self.minhours)
      self.hours_rem  = 0 if self.satisfied else (self.minhours - cumhours)

    # truncate unless greedy == True
    if self.greedy == False:
      if self.mincourses > 0:
        self.satcourses = self.satcourses[:self.mincourses]
      if self.minhours > 0:
        templist = []
        cumhours = 0
        for course in self.satcourses:
          templist.append(course)
          cumhours += course.credits
          if cumhours > self.minhours:
            break
        self.satcourses = templist

    # one more re-sort
    self.satcourses.sort(key = lambda course: (course.term, course.code))

    return 





class Group:
  '''
  Group()
  -------------
  A Group is the main organizing class for degrees, in that a Degree is simply a list of Groups.  In turn, a Group is a list of Requirements, supplemented by an (optional) list of Verifications.  The only difference between a Requirement and a Verification is that a Verification is not exclusive -- i.e. a single class can work toward satisfying multiple Verifications, whereas it can only satisfy a single Requirement within a Degree.
  '''


  def __init__(self, name=None, reqlist=None, verlist=None):

    self.name       = name
    self.reqlist    = [] if reqlist == None else reqlist
    self.verlist    = [] if verlist == None else verlist

    self.satisfied  = False
    self.satcourses = []
    self.hours_rem = None
    return 


  def add_requirement(self, name, courselist, mincourses=0, minhours=0, greedy=False):
    req = Requirement(name, courselist, mincourses, minhours, greedy)
    self.reqlist.append(req)
    return


  def add_verification(self, name, courselist, mincourses=0, minhours=0):
    req = Requirement(name, courselist, mincourses, minhours)
    self.verlist.append(req)
    return


  def check(self, coursehistory):

    for req in self.reqlist:
      req.check(coursehistory, self.verlist)
      for course in req.satcourses:
        self.satcourses.append(course)

    for ver in self.verlist:
      ver.check(coursehistory)
      #for course in ver.satcourses:        # Here is where a verifcation is 
      #  self.satcourses.append(course)     # different than a requirement

    reqs_satlist = [req.satisfied for req in self.reqlist]
    vers_satlist = [ver.satisfied for ver in self.verlist]
    self.satisfied = (False not in reqs_satlist and False not in vers_satlist)
    self.hours_rem = 0 if self.satisfied else sum([req.hours_rem for req in self.reqlist])
    return





class Degree:
   # ------------------------------
   # A Degree is a list of Groups.
   # ------------------------------


  def __init__(self, name=None, grplist=None):

    self.name = name
    self.grplist = [] if grplist == None else grplist
    self.complete = False
    self.required_credits = 0
    self.hours_rem = None

    # optional opportunity to customize degree checking and warning reporting
    self.pre_check_routines = []
    self.warning_checks_general = []
    self.warning_checks_major = []
    self.warnings_general = []
    self.warnings_major = []

    # add standard checks
    self.warning_checks_general.append(check_dropfails_lastyear)
    self.warning_checks_general.append(check_GPA_lastsemester)
    self.warning_checks_general.append(check_GPA_decrease)
    self.warning_checks_general.append(check_degree_GPA)


  def add_group(self, group):
    self.grplist.append(group)
    return

  def add_major_precheck(self, callback):
    self.pre_check_routines.append(callback)

  def add_major_warning_check(self, callback):
    self.warning_checks_major.append(callback)


  def check(self, student, coursehistory):

    # run any pre-check routines
    for pcr in self.pre_check_routines:
      pcr(student, coursehistory, self)


    # make a copy of the courselist and analyze it
    courses = copy.deepcopy(coursehistory)
    for grp in self.grplist:
      grp.check(courses)
      for rr in grp.satcourses:
        courses = [cc for cc in courses if cc.code != rr.code]  # removes already-used courses
    grp_satlist = [grp.satisfied for grp in self.grplist]
    self.complete = (False not in grp_satlist)
    if self.complete:
      self.hours_rem = 0
    else:
      self.hours_rem = sum([grp.hours_rem for grp in self.grplist])


    ## run any red-flag checks
    for wc in self.warning_checks_general:
      warning = wc(student, coursehistory, self)
      if warning != None:
        self.warnings_general.append(warning)

    for wc in self.warning_checks_major:
      warning = wc(student, coursehistory, self)
      if warning != None:
        self.warnings_major.append(warning)

    return




# -------------------------------
#    DPR Warnings
# -------------------------------

'''
The idea here is to abstract the idea of a DPR warning, so that red-flag conditions can be specified separately from the basic definitions of Degree / Group / Requirement / Course.  Crucially, it would be nice to have some "standard" or "general" red flags, but also allow departments to implement their own red-flag conditions in the major definition file.  

Not implemented yet -- currently red flags are caught in the class called ProgressReport() in this file.
'''


class DPRWarning:

  def __init__(self, level, message):
    self.level   = level
    self.message = message




def check_dropfails_lastyear(student, coursehistory, degree, threshhold=2.0):

    this_term = int(current_term())
    last_term = this_term - 5
    last_year = this_term - 10
    lycourses = [cc for cc in coursehistory if int(cc.term) >= last_year]

    # count dropped or failed courses last semester
    failset = set(['F'])
    weakset = set(['D-', 'D', 'D+'])
    dropset = set(['W', 'WF', 'WP', 'I'])
    fails = len([cc for cc in lycourses if cc.grade in failset])
    weaks = len([cc for cc in lycourses if cc.grade in weakset])
    drops = len([cc for cc in lycourses if cc.grade in dropset])
    score   = (3*fails + 2*weaks + 1*drops) / 2

    if score > 12:
      return DPRWarning(4, "very large number of drops/fails in past year")
    if score > 9:
      return DPRWarning(4, "large number of drops/fails in past year")
    if score > 6:
      return DPRWarning(4, "medium number of drops/fails in past year")
    if score > 3:
      return DPRWarning(4, "small number of drops/fails in past year")
    return None



def check_GPA_lastsemester(student, coursehistory, degree):

    this_term = int(current_term())

    # get last semester's courses
    last_term = this_term - 5
    courses_ls = [cc for cc in coursehistory if int(cc.term) == last_term and cc.grade in GPAgrades]

    # calculate last semester's GPA
    gpoints = np.array([GPAlookup[cc.grade] for cc in courses_ls]) 
    credits = np.array([cc.credits for cc in courses_ls])
    last_GPA = np.dot(gpoints, credits) / sum(credits)

    # return the appropriate response
    if last_GPA < 1.0:
      return DPRWarning(4, "low GPA last semester (%2.2f)" % (last_GPA))
    if last_GPA < 1.5:
      return DPRWarning(3, "low GPA last semester (%2.2f)" % (last_GPA))
    if last_GPA < 2.0:
      return DPRWarning(2, "low GPA last semester (%2.2f)" % (last_GPA))
    if last_GPA < 2.5:
      return DPRWarning(1, "low GPA last semester (%2.2f)" % (last_GPA))
    return None




def check_GPA_decrease(student, coursehistory, degree, threshhold=2.0):

    this_term = int(current_term())

    # get last semester's GPA
    last_term1 = this_term - 5
    courses_l1 = [cc for cc in coursehistory if int(cc.term) == last_term1 and cc.grade in GPAgrades]
    gpoints1 = np.array([GPAlookup[cc.grade] for cc in courses_l1]) 
    credits1 = np.array([cc.credits for cc in courses_l1])
    last_GPA1 = np.dot(gpoints1, credits1) / sum(credits1)

    # get two semester's past GPA
    last_term2 = this_term - 10
    courses_l2 = [cc for cc in coursehistory if int(cc.term) == last_term2 and cc.grade in GPAgrades]
    gpoints2 = np.array([GPAlookup[cc.grade] for cc in courses_l2]) 
    credits2 = np.array([cc.credits for cc in courses_l2])
    last_GPA2 = np.dot(gpoints2, credits2) / sum(credits2)

    decline = last_GPA2 - last_GPA1

    # return the appropriate response
    if decline > 2.0:
      return DPRWarning(4, "v. large GPA decline (%2.2f --> %2.2f)" % (last_GPA2, last_GPA1))
    if decline > 1.5:
      return DPRWarning(3, "large GPA decline (%2.2f --> %2.2f)" % (last_GPA2, last_GPA1))
    if decline > 1.0:
      return DPRWarning(2, "medium GPA decline (%2.2f --> %2.2f)" % (last_GPA2, last_GPA1))
    if decline > 0.5:
      return DPRWarning(1, "small GPA decline (%2.2f --> %2.2f)" % (last_GPA2, last_GPA1))
    return None



def check_degree_GPA(student, coursehistory, degree, threshhold=2.0):

    # get all degree courses
    degcourses = []
    for grp in degree.grplist:
      for cc in grp.satcourses:
        degcourses.append(cc)

    # calculate degree GPA
    degcourseset = set([cc.code for cc in degcourses])
    alldegcourses = [cc for cc in coursehistory if cc.code in degcourseset and cc.grade in GPAgrades]
    gpoints = np.array([GPAlookup[cc.grade] for cc in alldegcourses]) 
    credits = np.array([cc.credits for cc in alldegcourses])
    degree_GPA = np.dot(gpoints, credits) / sum(credits)
  
    # return the appropriate response
    if degree_GPA < 1.66:
      return DPRWarning(4, "prohibitively low degree GPA: %2.2f" % (degree_GPA))
    if degree_GPA < 2.00:
      return DPRWarning(3, "dangerously low degree GPA: %2.2f" % (degree_GPA))
    if degree_GPA < 2.33:
      return DPRWarning(2, "very low degree GPA: %2.2f" % (degree_GPA))
    if degree_GPA < 2.66:
      return DPRWarning(1, "low degree GPA: %2.2f" % (degree_GPA))
    return None






# ================================================
#     reporting and visualization
# ================================================


def terminal_courselist(student, courselist, degree):

  sortedcourses = copy.deepcopy(courselist)
  sortedcourses.sort(key=lambda course: "%s-%s" % (course.dept, course.code))
  for cc in sortedcourses:
    if cc.credits == 0: 
      continue
    print("{0:4} {1:4} --> {2:16} --> {3:4} {4:4}".format(cc.dept, cc.number, termcode_to_text(cc.term), cc.credits, cc.grade))
  print('\n')







def terminal_report(student, courselist, degree):
  '''
  terminal_report()
  -----------------
  This prints a report of the student's progress to the terminal, using Colorama for colored text.
  '''


  if degree.complete:
    dstatus = Fore.GREEN+"Complete!"+Style.RESET_ALL
  else:
    dstatus = Fore.RED+"Incomplete."+Style.RESET_ALL


  ois = outside_interests(student, courselist)
  oistring = ""
  for oi in ois:
    oistring += "%s(%d), " % (oi[0], oi[1])
  oistring = oistring[:-2]

  # student and degree information
  report  = ""
  report += student.lname + ", " + student.fname + "\n"
  report += "-"*25 + "\n"
  report += "SMU ID:           " + str(student.ID) + "\n"
  report += "Email:            " + str(student.email) + "\n"
  report += "Admitted:         " + termcode_to_text(student.aterm) + "\n"
  report += "Graduates:        " + termcode_to_text(student.gterm) + "\n"
  report += "GradStatus:       " + str(student.gstatus) + "\n"
  report += "\n"
  report += "Degree Type:      " + str(degree.name) + "\n"
  report += "Degree Status:    " + dstatus + "\n"
  report += "Other Interests:  " + oistring + "\n"
  report += "\n\n"

  # what to print for missing courses
  nullcourse = Course()
  nullcourse.term = '(none)'
  nullcourse.code = '(none)'
  nullcourse.grade = '-'

  # loop through the groups
  for grp in degree.grplist:

    grpstatus = Fore.GREEN+'satisfied'+Style.RESET_ALL if grp.satisfied else Fore.RED+'not satisfied'+Style.RESET_ALL
    report += "{0:24}: {1:16}".format(grp.name, grpstatus) + '\n'
    report += '-'*64 + '\n'

    # now loop through the requirements
    for req in grp.reqlist:
      counter = 0

      if (req.minhours > 0):
        sathours = sum([cc.credits for cc in req.satcourses])
        remhours = max(0, req.minhours - sathours)
        for kk,cc in enumerate(req.satcourses):
          counter = '' if (len(req.satcourses)==1 and remhours==0) else ' ' + str(kk+1)
          report += "{0:24} {1:16} {2:16} {3:4}".format(req.name+counter, cc.code, termcode_to_text(cc.term), cc.grade) + '\n'
        blanklns = int(np.ceil(remhours / 3.0))
        cc = nullcourse
        for kk in range(blanklns):
          report += "{0:24} {1:16} {2:16} {3:4}".format(req.name, cc.code, termcode_to_text(cc.term), cc.grade) + '\n'

      if (req.mincourses > 0):
        for kk,cc in enumerate(req.satcourses):
          counter = '' if req.mincourses == 1 else ' ' + str(kk+1)
          report += "{0:24} {1:16} {2:16} {3:4}".format(req.name+counter, cc.code, termcode_to_text(cc.term), cc.grade) + '\n'
        blanklns = req.mincourses - len(req.satcourses)
        cc = nullcourse
        for kk in range(blanklns):
          report += "{0:24} {1:16} {2:16} {3:4}".format(req.name, cc.code, termcode_to_text(cc.term), cc.grade) + '\n'

    # now loop through verifications
    for ver in grp.verlist:
      reqstatus = Fore.GREEN+'satisfied'+Style.RESET_ALL if ver.satisfied else Fore.RED+'not satisfied'+Style.RESET_ALL
      report += '{0:24}: {1}'.format(ver.name, reqstatus) + '\n'
    report += '\n\n'


  # General Warnings
  if len(degree.warnings_general) > 0:
    #report += 'General Warnings\n'
    #report += '--------------------\n'
    for w in degree.warnings_general:
      if w.level == 1:
        prefix = Fore.CYAN    + "Caution:  " + Style.RESET_ALL
      if w.level == 2:
        prefix = Fore.YELLOW  + "Concern:  " + Style.RESET_ALL
      if w.level == 3:
        prefix = Fore.RED     + "Warning:  " + Style.RESET_ALL
      if w.level == 4:
        prefix = Fore.MAGENTA + "PROBLEM:  " + Style.RESET_ALL
      report += prefix + w.message + ".\n"


  # Degree-Specific Warnings
  if len(degree.warnings_major) > 0:
    #report += 'Degree Plan Warnings\n'
    #report += '--------------------\n'
    for w in degree.warnings_major:
      if w.level == 1:
        prefix = Fore.CYAN    + "Caution:  " + Style.RESET_ALL
      if w.level == 2:
        prefix = Fore.YELLOW  + "Concern:  " + Style.RESET_ALL
      if w.level == 3:
        prefix = Fore.RED     + "Warning:  " + Style.RESET_ALL
      if w.level == 4:
        prefix = Fore.MAGENTA + "PROBLEM:  " + Style.RESET_ALL
      report += prefix + w.message + ".\n"

  report += '\n\n'
  return report





def pdf_report(student, courselist, degree, pdfname):
  '''
  pdf_report()
  -----------------
  This saves a textual report of the student's progress to the file $ID.pdf.
  '''

  if degree.complete:
    dstatus = "Complete!"
  else:
    dstatus = "Incomplete."

  ois = outside_interests(student, courselist)
  oistring = ""
  for oi in ois:
    oistring += "%s(%d), " % (oi[0], oi[1])
  oistring = oistring[:-2]

  # student and degree information
  report  = ""
  report += student.lname + ", " + student.fname + "\n"
  report += "-"*25 + "\n"
  report += "SMU ID:           " + str(student.ID) + "\n"
  report += "Email:            " + str(student.email) + "\n"
  report += "Admitted:         " + termcode_to_text(student.aterm) + "\n"
  report += "Graduates:        " + termcode_to_text(student.gterm) + "\n"
  report += "GradStatus:       " + str(student.gstatus) + "\n"
  report += "Degree Type:      " + str(degree.name) + "\n"
  report += "Degree Status:    " + dstatus + "\n"
  report += "Other Interests:  " + oistring + "\n"
  report += "\n\n"

  # loop through the groups
  for grp in degree.grplist:

    grpstatus = 'SATISFIED' if grp.satisfied else 'NOT SATISFIED'
    report += "{0:24}: {1:16}".format(grp.name, grpstatus) + '\n'
    report += '-'*64 + '\n'

    # what to print for missing courses
    nullcourse = Course()
    nullcourse.term = '________________'
    nullcourse.code = '________________'
    nullcourse.grade = '____'
    nc = nullcourse

    # now loop through the requirements
    for req in grp.reqlist:
      counter = 0

      if (req.minhours > 0):
        sathours = sum([cc.credits for cc in req.satcourses])
        remhours = max(0, req.minhours - sathours)
        for kk,cc in enumerate(req.satcourses):
          counter = '' if (len(req.satcourses)==1 and remhours==0) else ' ' + str(kk+1)
          report += "{0:24} {1:16} {2:16} {3:4}".format(req.name+counter, cc.code, termcode_to_text(cc.term), cc.grade) + '\n'
        blanklns = int(np.ceil(remhours / 3.0))
        cc = nullcourse
        for kk in range(blanklns):
          report += "{0:24} {1:16} {2:16} {3:4}".format(req.name, cc.code, termcode_to_text(cc.term), cc.grade) + '\n'

      if (req.mincourses > 0):
        for kk,cc in enumerate(req.satcourses):
          counter = '' if req.mincourses == 1 else ' ' + str(kk+1)
          report += "{0:24} {1:16} {2:16} {3:4}".format(req.name+counter, cc.code, termcode_to_text(cc.term), cc.grade) + '\n'
        blanklns = req.mincourses - len(req.satcourses)
        cc = nullcourse
        for kk in range(blanklns):
          report += "{0:24} {1:16} {2:16} {3:4}".format(req.name, cc.code, termcode_to_text(cc.term), cc.grade) + '\n'

    # now loop through verifications
    for ver in grp.verlist:
      reqstatus = 'SATISFIED' if ver.satisfied else 'NOT SATISFIED'
      report += '{0:24} {1}'.format(ver.name, reqstatus) + '\n'
    report += '\n\n'

  report += '\n'

  # General Warnings
  if len(degree.warnings_general) > 0:
    #report += 'General Warnings\n'
    #report += '--------------------\n'
    for w in degree.warnings_general:
      if w.level == 1:
        prefix = "Caution:  "
      if w.level == 2:
        prefix = "Concern:  "
      if w.level == 3:
        prefix = "Warning:  "
      if w.level == 4:
        prefix = "PROBLEM:  "
      report += prefix + w.message + ".\n"


  # Degree-Specific Warnings
  if len(degree.warnings_major) > 0:
    #report += 'Degree Plan Warnings\n'
    #report += '--------------------\n'
    for w in degree.warnings_major:
      if w.level == 1:
        prefix = "Caution:  "
      if w.level == 2:
        prefix = "Concern:  "
      if w.level == 3:
        prefix = "Warning:  "
      if w.level == 4:
        prefix = "PROBLEM:  "
      report += prefix + w.message + ".\n"

  # write to textfile
  sp = 'temp-report-%s' % (student.ID)
  outfile = open('%s.txt' % (sp), 'w')
  outfile.write(report)
  outfile.close()

  # convert to pdf and clean up
  call('enscript -p %s.ps --margins=72:72:72:72 --no-header %s.txt'%(sp, sp), shell=True)
  call('ps2pdf %s.ps %s'%(sp, pdfname), shell=True)
  call('rm %s.txt %s.ps'%(sp, sp), shell=True)

  return 





def plot_gpa_progression(courselist):
  '''
  Uses Matplotlib to bring up a plot of the student's GPA progression.
  '''

  terms  = [c.term for c in courselist]
  uterms = np.unique(terms)
  sterms = np.sort(uterms)
  aterms = (np.array( [float(t) for t in sterms] ) - 1000) / 10.0 + .05

  credits = np.empty(len(sterms))
  gpoints = np.empty(len(sterms))
  GPA  = np.empty(len(sterms))
  for kk,term in enumerate(sterms):
    courses     = [c for c in courselist if (c.term == term and c.grade in GPAgrades)]
    credits[kk] = sum([c.credits for c in courses])
    gpoints[kk] = sum([c.credits*GPAlookup[c.grade] for c in courses])
    GPA[kk]     = gpoints[kk] / credits[kk] if credits[kk] > 0 else 4.0

  totalGPA = sum(gpoints) / sum(credits)


  #trend lind for aterms, GPA, weighted by credits
  from lmfit import minimize, Parameters
  def model(params, x):
    value = params['a']
    slope = params['b']
    return value + slope * x

  def residual(params, x, data, weights):
    delta = data - model(params, x)
    return delta*weights

  params = Parameters()
  params.add('a', value=3.5)
  params.add('b', value=0.0)
  output = minimize(residual, params, args=(aterms, GPA, credits))


  fig, ax = plt.subplots(figsize=(6,4))

  # background
  ax.axhspan(0.00, 1.00, alpha=0.5, color='magenta')
  ax.axhspan(1.00, 1.66, alpha=0.5, color='red')
  ax.axhspan(1.66, 2.33, alpha=0.5, color='orange')
  ax.axhspan(2.33, 3.00, alpha=0.5, color='yellow')
  ax.axhspan(3.00, 3.66, alpha=0.5, color='greenyellow')
  ax.axhspan(3.67, 4.33, alpha=0.5, color='green')

  # GPA circles and average value
  ax.scatter(aterms, GPA, s=25*credits, color='white', marker='o', alpha=0.66, zorder=5)
  ax.plot(aterms, model(output.params, aterms), 'k--', linewidth=4, zorder=6)

  # setting axes
  xmin = np.floor(min(aterms))
  xmax = np.ceil(max(aterms))
  plt.xlim(xmin, xmax)
  plt.xticks(np.arange(xmin, xmax+1))
  myylim = min(1.33, min(GPA))
  plt.ylim(myylim - .33, 4.33)
  plt.title('Overall GPA is %1.3f' % (totalGPA))
  plt.xlabel('Term')
  plt.tight_layout()
  plt.show()
